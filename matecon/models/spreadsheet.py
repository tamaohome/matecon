from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path
from pprint import pformat
from typing import Final

from anytree import AnyNode
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.cell.read_only import EmptyCell, ReadOnlyCell
from openpyxl.worksheet.worksheet import Worksheet

from matecon.io.spreadsheet_reader import SpreadsheetReader
from matecon.models.position import Position
from matecon.utils.strings import is_valid_chars, zen2han

CellValueType = str | int | float
CellType = Cell | ReadOnlyCell | MergedCell | EmptyCell


class BookNode(AnyNode):
    """Excelブックを保持するノードクラス"""

    def __init__(self, filepath: str | Path, table_header: tuple[str, ...]):
        super().__init__(parent=None)
        self.filepath: Final[Path] = Path(filepath)
        self._table_header = table_header
        self._table: tuple[tuple, ...] = ()

        self._reader = SpreadsheetReader(self.filepath)
        for sheet in self._reader.get_worksheets():
            SheetNode(self, sheet)

    def __getitem__(self, key: int | str) -> SheetNode:
        if isinstance(index := key, int):
            if index < 0:
                index += len(self.sheets)
            if 0 <= index < len(self.sheets):
                return self.sheets[index]
            else:
                raise IndexError(f"BookNode インデックス '{index}' が範囲外です")

        if isinstance(name := key, str):
            for sheet in self:
                if sheet.name == name.strip():
                    return sheet
            raise KeyError(f"シート名 '{name}' が見つかりません")

        raise TypeError("キーには int 型または str 型を指定してください")

    def __len__(self):
        return len(self.sheets)

    def __iter__(self) -> Iterator[SheetNode]:
        return iter(self.sheets)

    def close(self) -> None:
        return self._reader.close()

    @property
    def name(self) -> str:
        """ファイル名（拡張子を含まない）"""
        return self.filepath.stem

    @property
    def filename(self) -> str:
        """ファイル名（拡張子を含む）"""
        return self.filepath.name

    @property
    def sheets(self) -> tuple[SheetNode]:
        """Excelシートのリストを返す"""
        return self.children

    @property
    def table(self) -> tuple[tuple, ...]:
        """ブック全体から抽出したテーブルを2次元リストとして返す"""
        if self._table == ():
            self._table = tuple(row for sheet in self.sheets for row in sheet.table)

        if self._table == ():
            raise ValueError(f"{self.filename}: 有効なシートが存在しません。")

        return self._table

    @property
    def table_header(self) -> tuple:
        """テーブルヘッダー"""
        return self._table_header

    @property
    def iter_table(self) -> Iterator[tuple]:
        return iter(self.table)


class SheetNode(AnyNode):
    """Excelシートを保持するノードクラス"""

    def __init__(self, parent: BookNode, worksheet: Worksheet):
        super().__init__(parent=parent)
        self._worksheet: Worksheet = worksheet
        self._table = self._get_table()

        # テーブル生成に失敗した場合、自ノードを除去
        if self._table == ():
            self.parent = None

    def __str__(self):
        return pformat(self.table)[1:-1]

    def __len__(self):
        return len(self.table)

    def __getitem__(self, index: int) -> tuple[CellValueType, ...]:
        if index < 0:
            index += len(self.table)
        if 0 <= index < len(self.table):
            return self.table[index]
        else:
            raise IndexError(f"SheetNode インデックス '{index}' が範囲外です")

    def __iter__(self):
        return iter(self.table)

    def cell(self, row_n: int, col_n: int) -> CellValueType | None:
        """セルの値を返す"""
        cell = self._worksheet.cell(row=row_n, column=col_n)
        if isinstance(cell, MergedCell):
            return None
        if isinstance(cell.value, CellValueType):
            return cell.value
        return None

    def iter_table(self) -> Iterator[tuple]:
        return iter(self.table)

    def _get_table(self) -> tuple[tuple, ...]:
        try:
            # テーブルの範囲を定義
            start = self.table_origin
            end = start + Position(0, len(self.table_header) - 1)

            # テーブルを生成
            iter_rows = self._worksheet.iter_rows(start.row, None, start.col, end.col)
            rows = [self._get_row(row) for row in iter_rows]
            return tuple(row for row in rows if any(row))

        except ValueError:
            return ()

    def _get_row(self, row: tuple) -> tuple:
        return tuple(cell.value for cell in row if isinstance(cell, CellType) and is_valid_chars(cell.value))

    def _formatted_cell_value(self, cell: CellType) -> CellValueType | None:
        """書式設定を考慮したセルの値を返す"""

        # セル値の型がその他の場合
        if not isinstance(cell.value, CellValueType):
            return str(cell.value)

        # セル値の型が文字列の場合
        if isinstance(cell.value, str):
            return cell.value

        # 結合されたセルの場合
        if isinstance(cell, MergedCell):
            return None

        # 空白セルの場合
        if isinstance(cell, EmptyCell):
            return None

        # セル値が整数の場合
        if isinstance(cell.value, int):
            return cell.value

        # セル値が小数を含む場合
        if isinstance(cell.value, float):
            return cell.value

        # TODO: 表示形式を適用
        # number_format = cell.number_format.replace("#", "0")
        # if cell.number_format == "0":
        #     return round(cell.value)

    @property
    def table(self) -> tuple[tuple, ...]:
        return self._table

    @property
    def name(self) -> str:
        return self._worksheet.title.strip()

    @property
    def booknode(self) -> BookNode:
        assert isinstance(self.parent, BookNode)
        return self.parent

    @property
    def table_header(self) -> tuple:
        return self.booknode.table_header

    @property
    def header_position(self) -> Position:
        """ヘッダー行の左端セルの位置を返す"""
        if self.table_header == ():
            raise ValueError("ヘッダーテンプレートが未定義です")

        for row_n, row in enumerate(self._worksheet.iter_rows(values_only=True), start=1):
            # ヘッダー行のセル値が全角英字の場合は半角英字に変換
            row = tuple(zen2han(c) for c in row)
            if set(self.table_header).issubset(row):
                for col_n, cell in enumerate(row, start=1):
                    if cell == self.table_header[0]:
                        return Position(row_n, col_n)

        raise ValueError(f"シート '{self.name}': ヘッダーが存在しません")

    @property
    def table_origin(self) -> Position:
        """テーブルの開始点"""
        return self.header_position + Position(1, 0)
