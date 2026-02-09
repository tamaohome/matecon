from __future__ import annotations

from contextlib import closing
from typing import Final

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.cell.read_only import EmptyCell
from openpyxl.worksheet.worksheet import Worksheet

from matecon.models.booknode import _SENTINEL, BookNode, SheetNode
from matecon.models.excel_file import ExcelFile
from matecon.models.templates import MATERIAL_HEADER

type CellType = str | int | float
type RowType = tuple[CellType | None, ...]
type SheetType = tuple[RowType, ...]


class ExcelReader:
    """Excelファイルを読み込むクラス"""

    def __init__(self, excel_file: ExcelFile, header=MATERIAL_HEADER):
        self.excel_file: Final = excel_file
        self.header: Final = header

    def load_booknode(self, ignore_hidden_sheet=True) -> BookNode:
        """Excelブックを取得する"""
        booknode = BookNode(self.excel_file, self.header, _SENTINEL)
        with closing(load_workbook(self.excel_file.filepath, read_only=True, data_only=True)) as wb:
            for worksheet in wb.worksheets:
                # 非表示シートの場合はスキップ
                if ignore_hidden_sheet and worksheet.sheet_state == "hidden":
                    continue
                self._set_sheetnode(booknode, worksheet)

        # 有効なシートが取得できなかった場合エラー
        if len(booknode) == 0:
            raise ValueError(f"{self.excel_file}: 有効なシートが存在しません。")

        return booknode

    def _set_sheetnode(self, booknode: BookNode, worksheet: Worksheet) -> None:
        """Excelシートを `BookNode` にセットする"""
        rows_iterator = worksheet.iter_rows(values_only=True)
        rows = tuple(self._get_row(row) for row in rows_iterator)
        SheetNode(booknode, worksheet.title, rows, _SENTINEL)

    def _get_row(self, row: tuple) -> RowType:
        return tuple(self._get_cell(cell) for cell in row)

    def _get_cell(self, cell) -> CellType | None:
        """書式設定を考慮したセルの値を返す"""
        if isinstance(cell, MergedCell):  # 結合されたセルの場合
            input(f"MergedCellを検出: {cell}")
            return None
        if isinstance(cell, EmptyCell) or cell is None:  # 空白セルの場合
            return None
        if isinstance(cell, int):  # 整数の場合
            return cell
        if isinstance(cell, float):  # 小数を含む場合
            return cell
        return str(cell)  # それ以外は文字列型として返す
